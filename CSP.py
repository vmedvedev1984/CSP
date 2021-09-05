from PyQt5 import QtWidgets, QtCore, QtGui, QtWidgets, uic
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QCheckBox
#from PyQt5.QtWidgets import QTableWidgetItem
#from PyQt5.QtGui import QPen, QColor, QImage, QPixmap, QPainter
#from PyQt5.QtCore import Qt, QTime, QCoreApplication, QEventLoop, QPointF
from openpyxl import Workbook, load_workbook
from docxtpl import DocxTemplate
import io, os, time, fnmatch, csv, datetime, json
from os import walk

ConfFileDialog = uic.loadUiType("conffile.ui")[0]

# ВОЗВРАЩАЕТ МАССИВ ПОДГОТОВЛЕННЫЙ ДЛЯ СНЯТИЯ ПОСЛЕДНЕГО DMT-КАТИОНА  use: lastDMT(oligolist).compare()
class lastDMT():
    def __init__(self, seqlist):
        self.seqlist = seqlist

    def compare(self):
        lenseq = 0
        for item in range(len(self.seqlist)):
            if len(self.seqlist[lenseq]) < len(self.seqlist[item]):
                lenseq = item
        for num in range(len(self.seqlist)):
            self.seqlist[num] = self.seqlist[num][0] + str( (len(self.seqlist[lenseq]) - len(self.seqlist[num])) * ' ') + self.seqlist[num][1:]
        return(self.seqlist)

#ОКНО НАСТРОЙКИ ПАРАМЕТРОВ СИНТЕЗА
class QDialogClass(QtWidgets.QDialog, ConfFileDialog):
    def __init__(self, parent=None):
        QtWidgets.QDialog.__init__(self, parent)
        self.setupUi(self)
        with open('config.txt') as json_file:  
            self.data = json.load(json_file)
        self.LoadDefault()
        self.pushButton_5.clicked.connect(self.AceptButton)
        self.pushButton_2.clicked.connect(self.dataFileButton)

    def LoadDefault(self):
        #Amidite
        self.doubleSpinBox.setValue(self.data['config'][1]['VolOneBase'])
        self.doubleSpinBox_7.setValue(self.data['config'][1]['DeadVol'])
        self.doubleSpinBox_38.setValue(self.data['config'][1]['MeCN'])
        self.doubleSpinBox_13.setValue(self.data['config'][1]['Amd'])
        #Oxidizer
        self.doubleSpinBox_6.setValue(self.data['config'][2]['VolOneBase'])
        self.doubleSpinBox_42.setValue(self.data['config'][2]['DeadVol'])
        self.doubleSpinBox_24.setValue(self.data['config'][2]['THF'])
        self.doubleSpinBox_25.setValue(self.data['config'][2]['Py'])
        self.doubleSpinBox_26.setValue(self.data['config'][2]['H2O'])
        self.doubleSpinBox_41.setValue(self.data['config'][2]['I2'])
        #Activator
        self.doubleSpinBox_2.setValue(self.data['config'][3]['VolOneBase'])
        self.doubleSpinBox_8.setValue(self.data['config'][3]['DeadVol'])
        self.doubleSpinBox_40.setValue(self.data['config'][3]['MeCN'])
        self.doubleSpinBox_14.setValue(self.data['config'][3]['TET'])
        #Deblock
        self.doubleSpinBox_3.setValue(self.data['config'][4]['VolOneBase'])
        self.doubleSpinBox_9.setValue(self.data['config'][4]['DeadVol'])
        self.doubleSpinBox_11.setValue(self.data['config'][4]['DCE'])
        self.doubleSpinBox_12.setValue(self.data['config'][4]['DCA'])
        #CapA
        self.doubleSpinBox_4.setValue(self.data['config'][5]['VolOneBase'])
        self.doubleSpinBox_10.setValue(self.data['config'][5]['DeadVol'])
        self.doubleSpinBox_15.setValue(self.data['config'][5]['THF'])
        self.doubleSpinBox_16.setValue(self.data['config'][5]['Anhydride'])
        #CapB
        self.doubleSpinBox_5.setValue(self.data['config'][6]['VolOneBase'])
        self.doubleSpinBox_39.setValue(self.data['config'][6]['DeadVol'])
        self.doubleSpinBox_19.setValue(self.data['config'][6]['THF'])
        self.doubleSpinBox_20.setValue(self.data['config'][6]['MeIm'])
        self.doubleSpinBox_21.setValue(self.data['config'][6]['Py'])
        #print(self.data['config'][9])
        self.textEdit.setText(str(self.data['config'][9]['protocolfile']))
        self.textEdit_2.setText(str(self.data['config'][8]['datafile']))
        self.textEdit_3.setText(str(self.data['config'][10]['omxdata']))

    
    def dataFileButton(self):
        self.dataFileName = QFileDialog.getExistingDirectory(self, 'Select a directory')
        self.textEdit_2.setText(str(self.dataFileName))

    
    def RejectButton(self):
        print('Reject')

    def AceptButton(self):
        self.data['config'][1]['VolOneBase'] = self.doubleSpinBox.value()
        self.data['config'][1]['DeadVol'] = self.doubleSpinBox_7.value()
        self.data['config'][1]['MeCN'] = self.doubleSpinBox_38.value()
        self.data['config'][1]['Amd'] = self.doubleSpinBox_13.value()
        #Oxidizer
        self.data['config'][2]['VolOneBase'] = self.doubleSpinBox_6.value()
        self.data['config'][2]['DeadVol'] = self.doubleSpinBox_42.value()
        self.data['config'][2]['THF'] = self.doubleSpinBox_24.value()
        self.data['config'][2]['Py'] = self.doubleSpinBox_25.value()
        self.data['config'][2]['H2O'] = self.doubleSpinBox_26.value()
        self.data['config'][2]['I2'] = self.doubleSpinBox_41.value()
        #Activator
        self.data['config'][3]['VolOneBase'] = self.doubleSpinBox_2.value()
        self.data['config'][3]['DeadVol'] = self.doubleSpinBox_8.value()
        self.data['config'][3]['MeCN'] = self.doubleSpinBox_40.value()
        self.data['config'][3]['TET'] = self.doubleSpinBox_14.value()
        #Deblock
        self.data['config'][4]['VolOneBase'] = self.doubleSpinBox_3.value()
        self.data['config'][4]['DeadVol'] = self.doubleSpinBox_9.value()
        self.data['config'][4]['DCE'] = self.doubleSpinBox_11.value()
        self.data['config'][4]['DCA'] = self.doubleSpinBox_12.value()
        #CapA
        self.data['config'][5]['VolOneBase'] = self.doubleSpinBox_4.value()
        self.data['config'][5]['DeadVol'] = self.doubleSpinBox_10.value()
        self.data['config'][5]['THF'] = self.doubleSpinBox_15.value()
        self.data['config'][5]['Anhydride'] = self.doubleSpinBox_16.value()
        #CapB
        self.data['config'][6]['VolOneBase'] = self.doubleSpinBox_5.value()
        self.data['config'][6]['DeadVol'] = self.doubleSpinBox_39.value()
        self.data['config'][6]['THF'] = self.doubleSpinBox_19.value()
        self.data['config'][6]['MeIm'] = self.doubleSpinBox_20.value()
        self.doubleSpinBox_21.setValue(self.data['config'][6]['Py'])
        #FilePath
        self.data['config'][9]['protocolfile'] = self.textEdit.toPlainText()
        self.data['config'][8]['datafile'] = self.textEdit_2.toPlainText()
        self.data['config'][10]['omxdata'] = self.textEdit_3.toPlainText()
        
        configfile = open("config.txt","w")
        configfile.writelines('{"config" : ' + str(json.dumps(self.data['config'], sort_keys=True, indent=4)) + '}')
        configfile.close()

class Window(QtWidgets.QMainWindow):
    def __init__(self):
        QtWidgets.QWidget.__init__(self)
        uic.loadUi("MainWindow.ui", self)
        #загрузка конфигурационного файла 
        self.fastalist = []
        with open('config.txt') as json_file:  
            self.data = json.load(json_file)
        #print(self.data)
        #self.label.setText("NewText")
        self.pushButton_6.clicked.connect(self.prepFa)
        self.pushButton_5.clicked.connect(self.stat)
        self.comboBox.addItems(["40nmol", "200nmol", "1umol"])
        self.comboBox_2.addItems(["ETT", "DCI"])
        self.pushButton.clicked.connect(self.postprocXLS)
        self.pushButton_2.clicked.connect(self.postXLSedit)
        #self.pushButton.clicked.connect(self.stat)
        self.tabWidget.setCurrentIndex(0)
        self.action_3.triggered.connect(self.conf_file_menu)

    def popupwin(self, text, title):
        infoBox = QMessageBox()
        infoBox.setIcon(QMessageBox.Information)
        infoBox.setText(text)
        infoBox.setWindowTitle(title)
        infoBox.setStandardButtons(QMessageBox.Ok)
        infoBox.exec_()        

    def conf_file_menu(self):
        dialog = QDialogClass()
        dialog.exec_()
        
    def stat(self):
        try:
            self.totemplate()
            wb = Workbook()
            #wb = openpyxl.load_workbook(filename = './template.xlsx')
            sheet = wb.active
            wb_first_row = ['Name', 'Sequence', 'Lenght', 'M.W.', 'Coef.Ext.', 'GC%', 'ug/OD250']
            for tab_colm in range(len(wb_first_row)):
                sheet.cell(row=1, column=tab_colm + 1).value = wb_first_row[tab_colm]
            for tab_row in range(2, len(self.fastalist) + 2):
                sheet.cell(row = tab_row, column = 1).value = self.tableWidget.item(tab_row - 2, 0).text()
                sheet.cell(row = tab_row, column = 2).value = self.tableWidget.item(tab_row - 2, 1).text()
                sheet.cell(row = tab_row, column = 3).value = int(self.tableWidget.item(tab_row - 2, 2).text())
                sheet.cell(row = tab_row, column = 4).value = self.tableWidget.item(tab_row - 2, 3).text().replace(".", ",")
                sheet.cell(row = tab_row, column = 5).value = self.tableWidget.item(tab_row - 2, 4).text()
                sheet.cell(row = tab_row, column = 6).value = self.tableWidget.item(tab_row - 2, 5).text().replace(".", ",")
                sheet.cell(row = tab_row, column = 7).value = self.tableWidget.item(tab_row - 2, 6).text().replace(".", ",")
            wb.save('./synthes/synthesis ' + str(datetime.date.today()) + '.xlsx')
            #ПЕРЕЗАПИСЬ OMXDATA ФАЙЛА
            omxdatafile = open("./omxdatatemp.txt","r+")
            oligolistnext =omxdatafile.readlines()
            for olig in range(2, len(self.fastalist) + 2):
                oligolistnext.append('Oligo:\t' + str(olig - 2) + '\t\t\t\t' + self.tableWidget.item(olig - 2, 1).text() +'\t' + str('Off' if int(self.tableWidget.cellWidget(olig-2, 7).checkState())==2 else 'On') + '\n')
            oligolistnext.append('##End##\t\t\t\t\t\t\n')
            omxdatafile = open("./omxdata.txt","r+")
            omxdatafile.truncate(0)
            omxdatafile.writelines(oligolistnext)
            omxdatafile.close()
            self.popupwin("Протокол синтеза готов", "Информация")
            '''if self.checkBox_3.isChecked():
                wb.save('./synthes/synthesis ' + str(datetime.date.today()) + '.xlsx')
            else:
                filenames = next(walk('./synthes/'), (None, None, []))[2]
                if(filenames[-1][9:-5]==str(datetime.date.today())):
                    wb.save('./synthes/synthesis ' + str(datetime.date.today()) + '-1.xlsx')
                else:
                    wb.save('./synthes/synthesis ' + str(datetime.date.today()) + '-' + str(int(filenames[-1][-6])+1) + '.xlsx')
                print(filenames)'''
        except ValueError:
                self.popupwin("Выберете файл с последовательностями олигонуклеотидов", 'Ошибка')
                return 0
        except AttributeError:
                self.popupwin("Выберете файл с последовательностями олигонуклеотидов", 'Ошибка')
                return 0    
    
    def totemplate(self):
        Mon_All = int(self.label_A_val.text())+int(self.label_C_val.text())+int(self.label_G_val.text())+int(self.label_T_val.text())
        context = { 'dA' : str(round((((float(self.label_A_val.text())*self.data['config'][1]['VolOneBase'])+self.data['config'][1]['DeadVol'])*(self.data['config'][1]['Amd']/self.data['config'][1]['MeCN'])),5)) + ' г',
                    'MeCN_dA' : str(round(((float(self.label_A_val.text())*self.data['config'][1]['VolOneBase'])+self.data['config'][1]['DeadVol']),4)) + ' мл',
                    'dC' : str(round((((float(self.label_C_val.text())*self.data['config'][1]['VolOneBase'])+self.data['config'][1]['DeadVol'])*(self.data['config'][1]['Amd']/self.data['config'][1]['MeCN'])),5)) + ' г',
                    'MeCN_dC' : str(round(((float(self.label_C_val.text())*self.data['config'][1]['VolOneBase'])+self.data['config'][1]['DeadVol']),4)) + ' мл',
                    'dG' : str(round((((float(self.label_G_val.text())*self.data['config'][1]['VolOneBase'])+self.data['config'][1]['DeadVol'])*(self.data['config'][1]['Amd']/self.data['config'][1]['MeCN'])),5)) + ' г',
                    'MeCN_dG' : str(round(((float(self.label_G_val.text())*self.data['config'][1]['VolOneBase'])+self.data['config'][1]['DeadVol']),4)) + ' мл',
                    'dT' : str(round((((float(self.label_T_val.text())*self.data['config'][1]['VolOneBase'])+self.data['config'][1]['DeadVol'])*(self.data['config'][1]['Amd']/self.data['config'][1]['MeCN'])),5)) + ' г',
                    'MeCN_dT' : str(round(((float(self.label_T_val.text())*self.data['config'][1]['VolOneBase'])+self.data['config'][1]['DeadVol']),4)) + ' мл',
                    'THF_Ox' : str(round(float(((Mon_All*self.data['config'][2]['VolOneBase'])+self.data['config'][2]['DeadVol'])*((self.data['config'][2]['THF'])/(self.data['config'][2]['THF']+self.data['config'][2]['Py']+self.data['config'][2]['H2O']))),5)) + ' мл',
                    'Py_Ox' : str(round(float(((Mon_All*self.data['config'][2]['VolOneBase'])+self.data['config'][2]['DeadVol'])*((self.data['config'][2]['Py'])/(self.data['config'][2]['THF']+self.data['config'][2]['Py']+self.data['config'][2]['H2O']))),5)) + ' мл',
                    'H2O_Ox' : str(round(float(((Mon_All*self.data['config'][2]['VolOneBase'])+self.data['config'][2]['DeadVol'])*((self.data['config'][2]['H2O'])/(self.data['config'][2]['THF']+self.data['config'][2]['Py']+self.data['config'][2]['H2O']))),5)) + ' мл',
                    'I2_Ox' : str(round(float(((Mon_All*self.data['config'][2]['VolOneBase'])+self.data['config'][2]['DeadVol'])*((self.data['config'][2]['I2'])/(self.data['config'][2]['THF']+self.data['config'][2]['Py']+self.data['config'][2]['H2O']))),5)) + ' г',
                    'V_Ox' : str(round(float((Mon_All*self.data['config'][2]['VolOneBase'])+self.data['config'][2]['DeadVol']),5)) + ' мл',
                    'MeCN_Act' : str(round(float((Mon_All*self.data['config'][3]['VolOneBase'])+self.data['config'][3]['DeadVol']),5)) + ' мл',
                    'TET_Act' : str(round(float(((Mon_All*self.data['config'][3]['VolOneBase'])+self.data['config'][3]['DeadVol'])*(self.data['config'][3]['TET']/self.data['config'][3]['MeCN'])),5)) + ' г',
                    'DCE_Dbl' : str(round(float(((Mon_All*self.data['config'][4]['VolOneBase'])+self.data['config'][4]['DeadVol'])*(self.data['config'][4]['DCE']/(self.data['config'][4]['DCE']+self.data['config'][4]['DCA']))),5)) + ' мл',
                    'DCA_Dbl' : str(round(float(((Mon_All*self.data['config'][4]['VolOneBase'])+self.data['config'][4]['DeadVol'])*(self.data['config'][4]['DCA']/(self.data['config'][4]['DCE']+self.data['config'][4]['DCA']))),5)) + ' мл',
                    'V_Dbl' : str(round(float((Mon_All*self.data['config'][4]['VolOneBase'])+self.data['config'][4]['DeadVol']),5)) + ' мл',
                    'THF_CPA' : str(round(float(((Mon_All*self.data['config'][5]['VolOneBase'])+self.data['config'][5]['DeadVol'])*(self.data['config'][5]['THF']/(self.data['config'][5]['THF']+self.data['config'][5]['Anhydride']))),5)) + ' мл',
                    'ANH_CPA' : str(round(float(((Mon_All*self.data['config'][5]['VolOneBase'])+self.data['config'][5]['DeadVol'])*(self.data['config'][5]['Anhydride']/(self.data['config'][5]['THF']+self.data['config'][5]['Anhydride']))),5)) + ' мл',
                    'V_CPA' : str(round(float((Mon_All*self.data['config'][5]['VolOneBase'])+self.data['config'][5]['DeadVol']),5)) + ' мл',
                    'MeCN_CPB' : str(round(float(((Mon_All*self.data['config'][6]['VolOneBase'])+self.data['config'][6]['DeadVol'])*(self.data['config'][6]['THF']/(self.data['config'][6]['THF']+self.data['config'][6]['MeIm']+self.data['config'][6]['Py']))),5)) + ' мл',
                    'MeIm_CPB' : str(round(float(((Mon_All*self.data['config'][6]['VolOneBase'])+self.data['config'][6]['DeadVol'])*(self.data['config'][6]['MeIm']/(self.data['config'][6]['THF']+self.data['config'][6]['MeIm']+self.data['config'][6]['Py']))),5)) + ' мл',
                    'Py_CPB' : str(round(float(((Mon_All*self.data['config'][6]['VolOneBase'])+self.data['config'][6]['DeadVol'])*(self.data['config'][6]['Py']/(self.data['config'][6]['THF']+self.data['config'][6]['MeIm']+self.data['config'][6]['Py']))),5)) + ' мл',
                    'V_CPB' : str(round(float((Mon_All*self.data['config'][6]['VolOneBase'])+self.data['config'][6]['DeadVol']),5)) + ' мл',
                    'DCI_Act' : str(round(float(((Mon_All*self.data['config'][7]['VolOneBase'])+self.data['config'][7]['DeadVol'])*(self.data['config'][7]['DCI']/self.data['config'][7]['MeCN'])),5)) + ' г',
                    'TET_Act_1' : True if self.comboBox_2.currentText() == 'ETT' else False,
                    'DCI_Act_1' : True if self.comboBox_2.currentText() == 'DCI' else False,
                    'month' : str(datetime.date.today()).split('-')[1],
                    'day' : str(datetime.date.today()).split('-')[2],
        }
        context.update({'oligos': [{'num' : str(y+1),
                                    'name' : str(self.tableWidget.item(y, 0).text() if self.tableWidget.item(y, 0) != None else ''),
                                    'seq' : str(self.tableWidget.item(y, 1).text() if self.tableWidget.item(y, 1) != None else ''),
                                    'base' : str(self.tableWidget.item(y, 2).text() if self.tableWidget.item(y, 2) != None else '')}  for y in range(len(self.fastalist))]})
        doc = DocxTemplate("protocol_template.docx")
        doc.render(context)
        doc.save("./protocols/Протокол синтеза олигонуклеотидов " + str(datetime.date.today()) + ".docx")
        #состояние CheckBox-ов DMT-On
        '''for dmt_check in range(len(self.fastalist)):
            print(int(self.tableWidget.cellWidget(dmt_check, 7).checkState()))'''
           

    # ЗАГРУЗКА ЭКСЕЛЬ ФАЙЛА ВО ВКЛАДКЕ ПОСТООБРАБОТКИ
    def postprocXLS(self):
        self.XLSdir = QFileDialog.getOpenFileName(None, 'Open File', './synthes/', "Excel Files (*.xlsx)")
        if self.XLSdir[0][-4:] == 'xlsx':
            self.textEdit_2.setText(self.XLSdir[0])
            wbCSV = load_workbook(filename = self.XLSdir[0])
            XLSopen = []
            sheet_ranges = wbCSV['Sheet']
            lensheet_range = 0
            while(lensheet_range < len(sheet_ranges['A'])):
                XLSopen.append(sheet_ranges['A' + str(lensheet_range + 1)].value)
                XLSopen.append(sheet_ranges['B' + str(lensheet_range + 1)].value)
                lensheet_range += 1
        else:
            return 0
        #print(XLSopen[2::2])
        for row in range(len(XLSopen[2::2])):
            self.tableWidget_2.insertRow(row)
            self.tableWidget_2.setItem(row, 0, QtWidgets.QTableWidgetItem(XLSopen[2*row + 2]))
            self.tableWidget_2.setItem(row, 1, QtWidgets.QTableWidgetItem(XLSopen[2*row + 3]))
        self.tableWidget_2.resizeColumnsToContents()
        
    # ЗАПОЛНЕНИЕ (РЕДАКТИРОВАНИЕ) ЭКСЕЛЬ ФАЙЛА ПРОТОКОЛА
    def postXLSedit(self):
        wbCSV = load_workbook(filename = self.XLSdir[0])
        sheet_ranges = wbCSV['Sheet']
        sheet_ranges.cell(row = 1, column = 8).value = "C, ng/ul"
        sheet_ranges.cell(row = 1, column = 9).value = "Add Water, ul"    
        for tab_r in range(len(sheet_ranges['A'])-1):
            try:
                sheet_ranges.cell(row = tab_r + 2, column = 8).value = self.tableWidget_2.item(tab_r, 2).text() if self.tableWidget_2.item(tab_r, 2).text() != None else ''
                sheet_ranges.cell(row = tab_r + 2, column = 9).value = (((float(self.tableWidget_2.item(tab_r, 2).text().replace(",", "."))) / float(sheet_ranges.cell(row = tab_r + 2, column = 4).value.replace(",", ".")))*1000)-100 if self.tableWidget_2.item(tab_r, 2).text() != None else 0
                self.tableWidget_2.setItem(tab_r, 3, QtWidgets.QTableWidgetItem(str(sheet_ranges.cell(row = tab_r + 2, column = 9).value)))
            except ValueError:
                self.popupwin("Заполните все ячейки концентрации числами", 'Ошибка')
                return 0
            except AttributeError:
                self.popupwin("Заполните все ячейки концентрации числами", 'Ошибка')
                return 0
        wbCSV.save(self.XLSdir[0])
        self.popupwin("Рабочий файл дополнен", "Информация")

    # ОТКРЫТИЕ ФАЙЛА ЗАКАЗА FASTA ИЛИ CSV И ВЫЧИСЛЕНИЕ И ФОРМИРОВАНИЕ ТАБЛИЦ НА ВКЛАДКЕ СИНТЕЗА
    def prepFa(self):
        self.fastadir = QFileDialog.getOpenFileName(None, 'Open File', './', "Fasta (*.fa *.fasta);;CSV Files (*.csv)")
        self.textEdit.setText(self.fastadir[0])
        if self.fastadir[0][-2:] == 'fa' or self.fastadir[0][-5:] == 'fasta':
            fastaopen = io.open(self.fastadir[0], mode='r').read().split()
            #print(fastaopen)
        elif self.fastadir[0][-3:] == 'csv':
            fastaopen = []
            f = open(self.fastadir[0], newline = '') 
            with f as csvfile:
                readCSV = csv.reader(csvfile, delimiter=';')
                for row in readCSV:
                    fastaopen.append('>' + row[0])
                    fastaopen.append(row[1])
            f.close()
        else:
            return 0
        i=0
        self.fastalist = []
        fastalistname = []
        coefex = []
        MWlist = []
        GClist =[]
        dA = 0
        dC = 0
        dG = 0
        dT = 0
        lenght_item = []
        while i<len(fastaopen):
            if fastaopen[i][0] == '>':
                fastalistname.append(fastaopen[i])
                i=i+1
            else:
                self.fastalist.append(fastaopen[i])
                #-------РАССЧЁТ КОЭФИЦИЕНТА ЭКСТИНКЦИИ ---------
                item_seq_ext = 0
                MW = 0
                GC = 0
                A = 0
                C = 0
                G = 0
                T = 0
                leight = 0
                for j in range(len(fastaopen[i])-1):
                    if fastaopen[i][j] == "A" :
                        dA = dA + 1 
                        A = A + 1
                        MW = MW + 251.2
                        if fastaopen[i][j+1] == "A":
                            item_seq_ext = item_seq_ext + 27400 - 15400  #ex(AA)-ex(A)
                        elif fastaopen[i][j+1] == "C" :
                            item_seq_ext = item_seq_ext + 21200 - 7400 #ex(AC)-ex(A)
                        elif fastaopen[i][j+1] == "G":
                            item_seq_ext = item_seq_ext + 25000 - 11500 #ex(AG)-ex(A)
                        elif fastaopen[i][j+1] == "T":
                            item_seq_ext = item_seq_ext + 22800 - 8700 #ex(AT)-ex(A)
                    if fastaopen[i][j] == "C" :
                        dC = dC +1
                        C = C + 1
                        MW = MW + 227.2
                        GC = GC + 1
                        if fastaopen[i][j+1] == "A":
                            item_seq_ext = item_seq_ext + 21200 - 15400 #ex(CA)-ex(A)
                        elif fastaopen[i][j+1] == "C" :
                            item_seq_ext = item_seq_ext + 14600 - 7400 #ex(CC)-ex(A)
                        elif fastaopen[i][j+1] == "G":
                            item_seq_ext = item_seq_ext + 18000 - 11500 #ex(CG)-ex(A)
                        elif fastaopen[i][j+1] == "T":
                            item_seq_ext = item_seq_ext + 15200 - 8700 #ex(CT)-ex(A)
                    if fastaopen[i][j] == "G" :
                        dG = dG +1
                        G = G + 1
                        MW = MW + 267.2
                        GC = GC + 1
                        if fastaopen[i][j+1] == "A" :
                            item_seq_ext = item_seq_ext + 25200 - 15400 #ex(GA)-ex(A)
                        elif fastaopen[i][j+1] == "C" :
                            item_seq_ext = item_seq_ext + 17600 - 7400 #ex(GC)-ex(A)
                        elif fastaopen[i][j+1] == "G":
                            item_seq_ext = item_seq_ext + 21600 - 11500 #ex(GG)-ex(A)
                        elif fastaopen[i][j+1] == "T":
                            item_seq_ext = item_seq_ext + 20000 - 8700 #ex(GT)-ex(A)
                    if fastaopen[i][j] == "T" :
                        dT = dT + 1
                        T = T + 1
                        MW = MW + 242.2
                        if fastaopen[i][j+1] == "A" :
                            item_seq_ext = item_seq_ext + 23400 - 15400 #ex(GA)-ex(A)
                        elif fastaopen[i][j+1] == "C" :
                            item_seq_ext = item_seq_ext + 16200 - 7400 #ex(GC)-ex(A)
                        elif fastaopen[i][j+1] == "G":
                            item_seq_ext = item_seq_ext + 19000 - 11500 #ex(GG)-ex(A)
                        elif fastaopen[i][j+1] == "T":
                            item_seq_ext = item_seq_ext + 16800 - 8700 #ex(GT)-ex(A)
                if fastaopen[i][-1] == "A":
                    dA = dA +1
                    A = A + 1
                    MW = MW + 251.2
                    item_seq_ext = item_seq_ext + 15400
                elif fastaopen[i][-1] == "C":
                    dC = dC + 1
                    C = C + 1
                    MW = MW +227.2
                    GC = GC + 1
                    item_seq_ext = item_seq_ext + 7400
                elif fastaopen[i][-1] == "G":
                    dG = dG +1
                    G = G + 1
                    MW = MW + 267
                    GC = GC + 1
                    item_seq_ext = item_seq_ext + 11500
                elif fastaopen[i][-1] == "T":
                    dT = dT + 1
                    T = T + 1
                    MW = MW + 242.2
                    item_seq_ext = item_seq_ext + 8700
                MW = MW + (len(fastaopen[i])-1)*62.05        #фосфатный остов
                GC = (GC / len(fastaopen[i]))*100
                coefex.append(item_seq_ext)
                MWlist.append(MW)
                GClist.append(GC)
                leight = A + T + C + G
                lenght_item.append(leight)
                i=i+1
        #print(self.fastalist) СДЕЛАТЬ ИСКЛЮЧЕНИЕ НА МОДИФИКАЦИИ
        # ЗАПОЛНЕНИЕ ТАБЛИЦЫ И ЗНАЧЕНИЙ КОЛИЧЕСТВА АМИДИТОВ НА ВЛАДКЕ СИНТЕЗА
        for row in range(len(self.fastalist)):
            self.tableWidget.insertRow(row)
            self.tableWidget.setItem(row, 0, QtWidgets.QTableWidgetItem(fastalistname[row][1:]))
            self.tableWidget.setItem(row, 1, QtWidgets.QTableWidgetItem(self.fastalist[row]))
            self.tableWidget.setItem(row, 2, QtWidgets.QTableWidgetItem(str(lenght_item[row])))
            self.tableWidget.setItem(row, 3, QtWidgets.QTableWidgetItem(str(format(MWlist[row], '.2f'))))
            self.tableWidget.setItem(row, 4, QtWidgets.QTableWidgetItem(str(coefex[row])))
            self.tableWidget.setItem(row, 5, QtWidgets.QTableWidgetItem(str(format(GClist[row], '.2f'))))
            self.tableWidget.setItem(row, 6, QtWidgets.QTableWidgetItem(str(format(((MWlist[row]/coefex[row])*1000), '.2f'))))
            #self.tableWidget.cellWidget(row, 7).checkState()
            self.tableWidget.setCellWidget(row,7, QCheckBox())
            #self.tableWidget.setItem(row, 7, QtWidgets.QTableWidgetItem(QCheckBox('Show title', self)))
            self.label_A_val.setText(str(dA))
            self.label_C_val.setText(str(dC))
            self.label_G_val.setText(str(dG))
            self.label_T_val.setText(str(dT))
        
        
        
        

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    w = Window()
    w.show()
    sys.exit(app.exec_())
